from django.shortcuts import render,redirect
from .forms import DiaryForm
from .models import Diary
import openpyxl
from openpyxl import Workbook
import os
from openpyxl import load_workbook

# Create your views here.

def diary_list(request):
    context = {'diary_list': Diary.objects.all()}
    return render(request, "goshenlechem_register/diary_list.html",context)

import openpyxl
from openpyxl import Workbook

def save_to_excel(data, update=False, identifier=None):
    # Path to your Excel file
    file_path = "C:\\Users\\USER\\Desktop\\CompanyContactDiary.xlsx"
    
    try:
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        # Write headers (optional)
        headers = list(data.keys())
        sheet.append(headers)
    
    if update and identifier:
        # Find the row that needs to be updated
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            if row[0].value == identifier:  # Assuming 'First Name' is the identifier
                for idx, key in enumerate(data.keys(), start=1):
                    row[idx-1].value = data[key]
                break
    else:
        # Append data as a new row
        row = list(data.values())
        sheet.append(row)
    
    # Save the workbook
    workbook.save(file_path)

def diary_form(request, id=0):
    if request.method == "GET":
        if id == 0:
            form = DiaryForm()
        else:
            diary = Diary.objects.get(pk=id)
            form = DiaryForm(instance=diary)
        return render(request, "goshenlechem_register/diary_form.html", {'form': form})
    else:
        if id == 0:
            form = DiaryForm(request.POST)
        else:
            diary = Diary.objects.get(pk=id)
            form = DiaryForm(request.POST, instance=diary)
        if form.is_valid():
            instance = form.save()
            # Prepare data for Excel
            data = {
                'First Name': instance.firstname, 
                'Second Name': instance.secondname,
                'Date Of Birth': instance.dateofbirth,
                'Company Of Operation': instance.companyofoperation,
                'Position': instance.position.title if instance.position else '',
                'Country': instance.country,
                'Contact': instance.contact,
                'Email': instance.email,
                # Add all relevant fields from the form/instance
            }
            # Determine whether to update or insert
            update = id != 0
            # Use the primary identifier (like 'First Name') to find the row to update
            identifier = instance.firstname
            # Save or update data in the Excel file
            save_to_excel(data, update=update, identifier=identifier)
        return redirect('/diary/list')

            
#Delete operations
def diary_delete(request, id):
    import os
    from openpyxl import load_workbook
    from django.shortcuts import redirect, get_object_or_404
    from .models import Diary

    # Fetch the diary entry safely
    diary = get_object_or_404(Diary, pk=id)

    # File path to the Excel file
    file_path = "C:\\Users\\USER\\Desktop\\CompanyContactDiary.xlsx"  # Adjust this path as needed

    # Delete the diary entry from the database
    diary.delete()

    # Now remove the corresponding entry from the Excel file
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active

        # Iterate through rows to find the matching row based on the details provided
        for row in sheet.iter_rows(min_row=2):  # Start from row 2 to skip the header
            if (
                str(row[0].value) == diary.firstname and
                str(row[1].value) == diary.secondname and
                str(row[2].value) == diary.dateofbirth and  # No need to convert, it's already a string
                str(row[3].value) == diary.companyofoperation and
                str(row[4].value) == diary.position and
                str(row[5].value) == diary.country and
                str(row[6].value) == diary.contact and
                str(row[7].value) == diary.email
            ):
                sheet.delete_rows(row[0].row)  # Delete the entire row
                break  # Exit the loop after deleting

        # Save the workbook after deletion
        wb.save(file_path)

    return redirect('/diary/list')

